'''
Assignment - Extracting financial news headline from google finance

Objective is to automate the process of extracting local financial news from the google finance site.

Steps
=====
 > Goto https://finance.google.com/
 > Goto news section and select local news
 > Scrape all the news with the links and financial ticker(s) mentioned below the news.
 > Store the values as expected by the user (template shared).

What is expected
=================
 > A workflow (flowchart) with the input, output, exceptions defined.
 > The automation script (using python)
 > A small presentation to explain the solution.
 > Video of the automation (screen capture)
 
'''

#____________________________________________________________________________________________________________
#____________________________________________________________________________________________________________

# Importing important libraries
import time
import openpyxl
from datetime import datetime
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


#____________________________________________________________________________________________________________
#____________________________________________________________________________________________________________

# A funtion to Save the data from website
def saving_to_xlsx(data_list):
    # Open a new workbook
    wb = openpyxl.Workbook()

    # Create a new sheet
    sheet = wb.active

    # Add column headers
    sheet.cell(row=1, column=1).value = "Headline"
    sheet.cell(row=1, column=2).value = "Tickers"
    sheet.cell(row=1, column=3).value = "Link"

    # Start adding data from row 2
    row_counter = 2
    for data in data_list:
        sheet.cell(row=row_counter, column=1).value = data["Headline"]
        sheet.cell(row=row_counter, column=1).alignment = Alignment(wrap_text=True)
        # Calculate optimal row height based on the wrapped text
        cell_row = sheet.row_dimensions[sheet.cell(row=row_counter, column=1).row]
        cell_row.auto_size = True
        
        sheet.cell(row=row_counter, column=2).value = data["Tickers"]
        sheet.cell(row=row_counter, column=2).alignment = Alignment(wrap_text=True)
        # Calculate optimal row height based on the wrapped text
        cell_row = sheet.row_dimensions[sheet.cell(row=row_counter, column=2).row]
        cell_row.auto_size = True
        
        
        sheet.cell(row=row_counter, column=3).value = data["Link"]
        sheet.cell(row=row_counter, column=3).alignment = Alignment(wrap_text=True)
        # Calculate optimal row height based on the wrapped text
        cell_row = sheet.row_dimensions[sheet.cell(row=row_counter, column=3).row]
        cell_row.auto_size = True
        
        row_counter += 1
    
    # Set column widths
    sheet.column_dimensions['A'].width = 50  # Headline
    sheet.column_dimensions['B'].width = 15   # Tickers
    sheet.column_dimensions['C'].width = 50 # Link
    

    # Saving the file name in date and time 
    current_datetime = datetime.now().strftime("%d-%m-%Y_%H:%M:%S")
    xlsx_file_path = f"./output_dataset/output_{current_datetime}.xlsx"

    # Save the workbook as an XLSX file
    wb.save(xlsx_file_path)

    print("Articles successfully saved to", xlsx_file_path)


#____________________________________________________________________________________________________________
#____________________________________________________________________________________________________________

# A funtion to Scarpe the data from website
def scrape_assignment_from_google_finance(url="https://www.google.com/finance/"):

    try:
        
        driver = webdriver.Chrome()  # driver for Chrome

        # Open the webpage
        driver.get(url)

        # Wait for the page to load (3 seconds)
        time.sleep(3)

        # Select local market tab
        try:
            local_market_tab = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '[data-tab-id="localMarketNews"]'))
            )
            local_market_tab.click()
        except (TimeoutException, NoSuchElementException):
            print("Local market tab not found or timed out.")
            return
        
        time.sleep(3)

        # Wait for articles section to load
        try:
            articles_container = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div[jsshadow][jscontroller="dOH8Ue"]'))
            )
        except (TimeoutException, NoSuchElementException):
            print("Articles section not found or timed out.")
            return

        # Extract articles
        article_elements = articles_container.find_elements(By.CLASS_NAME, "yY3Lee")

        if not article_elements:
            print("No articles found.")
            return

        data_list = []
        for article in article_elements:
            try:
                headline = article.find_element(By.CSS_SELECTOR, '.Yfwt5').text.strip()
                link = article.find_element(By.CSS_SELECTOR, 'a[data-ved][jslog][href]').get_attribute('href')
                tickers = ", ".join(
                    ticker.text.strip() for ticker in article.find_elements(By.CSS_SELECTOR, '.X18JZ')
                )
                data_list.append({"Headline": headline, "Tickers": tickers, "Link": link})
            except NoSuchElementException:
                print("Error extracting data from an article.")

        #print(data_list)
        
        
        print("Articles scraped successfully")

        xlsx_file_path = "output.xlsx"

        # Saving to xlsx file
        saving_to_xlsx(data_list)

        print("END OF THE PROGRAM")

    except (TimeoutException, NoSuchElementException) as e:
        print("An error occurred:", e)
    finally:
        driver.quit()

#____________________________________________________________________________________________________________
#____________________________________________________________________________________________________________
        
# Calling the function
scrape_assignment_from_google_finance()
  
